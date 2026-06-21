"""Generate static lookup JSON files from the Attribution Glossary xlsx.

Outputs (written into aggregator/ so they get bundled with the function):
  - incident_reason_map.json : INCIDENT_REASON code -> Incident Category Description
  - period_end_dates.json    : "YYYY/YY_Pn" -> ISO end date of that railway period
  - operator_map.json        : TOC_CODE -> {name, class}  (class = passenger|freight)

Run whenever the glossary changes (rarely):
    python scripts/build_lookups.py path/to/Glossary.xlsx
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = ROOT / "aggregator"

# Operators with no fare-paying passengers: freight, infrastructure, engineering,
# light-loco and charter-haulage companies. Their delay is bridged to "equivalent
# passenger-hours" via the TAG value-of-time ratio (see conversion_params.json)
# rather than multiplied by a passenger load. Matched against the glossary name.
_FREIGHT_KW = re.compile(
    r"freight|cargo|dbc|schenker|railfreight|freightliner|\bdrs\b|colas|loram|infra|"
    r"harsco|rail operations|gb rail eng|wcr|swietelsky|babcock|amey|volker|carillion|"
    r"balfour|network rail|serco rail ops|legge|seco|jsd|premet|railadventure|loco|"
    r"locomotive|victa|fastline|hanson|dcr|europorte|varamis|slc|railtest|rits|wrong tsc",
    re.I,
)
# Non-mainline operators that DO carry passengers (heritage / metro / charter /
# international) -> keep on the passenger side even if a freight keyword matches.
_PASSENGER_KW = re.compile(
    r"charter|vsoe|orient|pullman|vintage|steam|nymr|swanage|supertram|nexus|"
    r"eurostar|lul|underground",
    re.I,
)


def classify_operator(name):
    """passenger | freight, from the operator's glossary name."""
    name = (name or "").strip()
    if _PASSENGER_KW.search(name):
        return "passenger"
    if _FREIGHT_KW.search(name):
        return "freight"
    return "passenger"


def build_operator_map(wb):
    """'Operator Name' sheet: TOC code -> {name, class}."""
    ws = wb["Operator Name"]
    rows = list(ws.iter_rows(values_only=True))[1:]  # skip header
    out = {}
    for r in rows:
        code = r[0]
        if not code:
            continue
        name = str(r[1]).strip() if len(r) > 1 and r[1] else ""
        out[str(code).strip()] = {"name": name, "class": classify_operator(name)}
    return out


def build_reason_map(wb):
    ws = wb["Incident Reason"]
    rows = list(ws.iter_rows(values_only=True))
    hdr = rows[0]
    code_i = hdr.index("Incident Reason")
    desc_i = hdr.index("Incident Category Description")
    out = {}
    for r in rows[1:]:
        code = r[code_i]
        if not code:
            continue
        desc = (str(r[desc_i]).strip() if r[desc_i] else "")
        out[str(code).strip()] = desc
    return out


def build_period_end_dates(wb):
    """The 'Period Dates' sheet lays years out across column blocks of 5.

    Header row holds 'YEAR 2024/25'; within each block the columns are
    [Day Name, '', Date, '', No of Days]. Period rows run Period 1..13.
    """
    ws = wb["Period Dates"]
    rows = list(ws.iter_rows(values_only=True))

    # Find the row containing the 'YEAR ....' labels and the period-data rows.
    year_row_idx = None
    for i, r in enumerate(rows):
        if any(isinstance(c, str) and c.strip().startswith("YEAR ") for c in r):
            year_row_idx = i
            break
    if year_row_idx is None:
        return {}

    year_row = rows[year_row_idx]
    # Map the column index of each 'YEAR xxxx/yy' label -> the FY string.
    year_cols = {}
    for ci, c in enumerate(year_row):
        if isinstance(c, str) and c.strip().startswith("YEAR "):
            year_cols[ci] = c.strip().replace("YEAR ", "").strip()  # e.g. "2024/25"

    out = {}
    for r in rows[year_row_idx + 1:]:
        label = r[0]
        if not (isinstance(label, str) and label.strip().lower().startswith("period")):
            continue
        try:
            pnum = int(label.strip().split()[-1])
        except ValueError:
            continue
        for ci, fy in year_cols.items():
            date_cell = r[ci + 2] if ci + 2 < len(r) else None  # Date sits 2 cols right of label
            if date_cell is None:
                continue
            iso = None
            if hasattr(date_cell, "date"):
                iso = date_cell.date().isoformat()
            elif isinstance(date_cell, str) and date_cell.strip():
                iso = date_cell.strip()[:10]
            if iso:
                # Match the CSV's zero-padded FINANCIAL_YEAR_PERIOD, e.g. "2026/27_P01".
                out[f"{fy}_P{pnum:02d}"] = iso
    return out


def main():
    xlsx = sys.argv[1] if len(sys.argv) > 1 else str(
        Path.home() / "Downloads" / "Transparency page Attribution Glossary.xlsx"
    )
    wb = openpyxl.load_workbook(xlsx, data_only=True)

    reason_map = build_reason_map(wb)
    period_dates = build_period_end_dates(wb)
    operator_map = build_operator_map(wb)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    (OUT_DIR / "incident_reason_map.json").write_text(
        json.dumps(reason_map, indent=2, sort_keys=True)
    )
    (OUT_DIR / "period_end_dates.json").write_text(
        json.dumps(period_dates, indent=2, sort_keys=True)
    )
    (OUT_DIR / "operator_map.json").write_text(
        json.dumps(operator_map, indent=2, sort_keys=True)
    )
    n_frt = sum(1 for v in operator_map.values() if v["class"] == "freight")
    print(f"incident_reason_map.json : {len(reason_map)} codes")
    print(f"period_end_dates.json    : {len(period_dates)} periods")
    print(f"operator_map.json        : {len(operator_map)} operators "
          f"({n_frt} freight / {len(operator_map) - n_frt} passenger)")
    # quick sanity
    print("  sample:", "TN ->", reason_map.get("TN"),
          "| 2024/25_P11 end ->", period_dates.get("2024/25_P11"),
          "| WA ->", operator_map.get("WA"))


if __name__ == "__main__":
    main()
